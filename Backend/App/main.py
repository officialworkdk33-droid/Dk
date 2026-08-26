import io
import os
import uuid
from typing import Any

import httpx
from dotenv import load_dotenv
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from pypdf import PdfReader

load_dotenv()

app = FastAPI(title="Customs Invoice AI", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

BRIDGE_URL = os.getenv("BRIDGE_URL", "http://127.0.0.1:8787").rstrip("/")
BRIDGE_TOKEN = os.getenv("BRIDGE_TOKEN", "change-me")
MAX_FILE_MB = int(os.getenv("MAX_FILE_MB", "20"))

SYSTEM_PROMPT = """You are a customs invoice document extraction engine.
Extract only information supported by the supplied document text.
Never invent values. Use null for unavailable scalar fields and [] for unavailable lists.
Return ONLY valid JSON matching the requested schema.
For ambiguous fields, add a warning.
Preserve invoice values and units exactly where practical.
HS codes must be returned only when explicitly present or strongly identified in the document; do not guess.
Confidence is 0 to 1.
"""

SCHEMA = {
    "invoice_number": None,
    "invoice_date": None,
    "supplier_name": None,
    "supplier_address": None,
    "buyer_name": None,
    "buyer_address": None,
    "consignee": None,
    "notify_party": None,
    "country_of_origin": None,
    "country_of_destination": None,
    "currency": None,
    "incoterms": None,
    "purchase_order": None,
    "transport_mode": None,
    "airway_bill": None,
    "bill_of_lading": None,
    "total_quantity": None,
    "total_net_weight": None,
    "total_gross_weight": None,
    "total_invoice_value": None,
    "line_items": [],
    "confidence": 0,
    "warnings": [],
}

def extract_pdf_text(data: bytes) -> str:
    try:
        reader = PdfReader(io.BytesIO(data))
        return "\n\n".join((page.extract_text() or "") for page in reader.pages).strip()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Unable to read PDF: {exc}")

async def call_bridge(text: str, filename: str) -> dict[str, Any]:
    payload = {
        "filename": filename,
        "system_prompt": SYSTEM_PROMPT,
        "schema": SCHEMA,
        "text": text,
    }
    try:
        async with httpx.AsyncClient(timeout=180) as client:
            response = await client.post(
                f"{BRIDGE_URL}/extract",
                json=payload,
                headers={"Authorization": f"Bearer {BRIDGE_TOKEN}"},
            )
            response.raise_for_status()
            return response.json()["data"]
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=502, detail=f"AI bridge unavailable: {exc}")

def make_excel(result: dict[str, Any], filename: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    fields = [
        "invoice_number", "invoice_date", "supplier_name", "supplier_address",
        "buyer_name", "buyer_address", "consignee", "notify_party",
        "country_of_origin", "country_of_destination", "currency", "incoterms",
        "purchase_order", "transport_mode", "airway_bill", "bill_of_lading",
        "total_quantity", "total_net_weight", "total_gross_weight",
        "total_invoice_value", "confidence",
    ]
    ws.append(["Source File", filename])
    for field in fields:
        ws.append([field, result.get(field)])

    ws2 = wb.create_sheet("Line Items")
    headers = [
        "item_no", "description", "hs_code", "country_of_origin",
        "quantity", "unit", "unit_price", "amount", "currency",
        "net_weight", "gross_weight"
    ]
    ws2.append(headers)
    for item in result.get("line_items") or []:
        ws2.append([item.get(h) for h in headers])

    ws3 = wb.create_sheet("Warnings")
    ws3.append(["warning"])
    for warning in result.get("warnings") or []:
        ws3.append([warning])

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

@app.get("/health")
async def health():
    return {"ok": True, "service": "customs-invoice-ai"}

@app.post("/api/extract")
async def extract(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are supported.")
    data = await file.read()
    if len(data) > MAX_FILE_MB * 1024 * 1024:
        raise HTTPException(status_code=413, detail=f"PDF exceeds {MAX_FILE_MB} MB.")

    text = extract_pdf_text(data)
    if not text:
        return {
            "id": str(uuid.uuid4()),
            "filename": file.filename,
            "status": "ocr_required",
            "message": "No text layer was found. This PDF appears scanned/image-only. OCR is required.",
            "data": SCHEMA,
        }

    result = await call_bridge(text, file.filename)
    return {"id": str(uuid.uuid4()), "filename": file.filename, "status": "extracted", "data": result}

@app.post("/api/export")
async def export(payload: dict[str, Any]):
    result = payload.get("data")
    filename = payload.get("filename", "invoice")
    if not isinstance(result, dict):
        raise HTTPException(status_code=400, detail="Invalid extraction data.")
    xlsx = make_excel(result, filename)
    safe = filename.rsplit(".", 1)[0].replace(" ", "_")
    return StreamingResponse(
        io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe}_extracted.xlsx"'},
    )
